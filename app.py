from flask import Flask, request, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_marshmallow import Marshmallow
from flask_jwt_extended import (
    create_access_token,
    jwt_required,
    get_jwt_identity,
    JWTManager,
)

from openpyxl import load_workbook


app = Flask(__name__)

app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///database.db"
app.config["JWT_SECRET_KEY"] = "my-jwt-secret-key"
app.config["SECRET_KEY"] = "my-secret-key"
app.config["JWT_ACCESS_TOKEN_EXPIRES"] = False  # token will not expire here...

db = SQLAlchemy(app)
ma = Marshmallow(app)
jwt = JWTManager(app)


class Parser(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50))
    age = db.Column(db.String(30))
    location = db.Column(db.String(50))
    username = db.Column(db.String(50))
    password = db.Column(db.String(30))


class ParserSchema(ma.Schema):
    class Meta:
        fields = ["id", "name", "age", "location", "username", "password"]


Parser_schema = ParserSchema()
Parser_schemas = ParserSchema(many=True)

#  Authentication JWT...

# register user..

users = []


"""@app.route("/register", methods=["POST"])
def register():
    data = request.get_json()
    username = data["username"]
    password = data["password"]

    if not username or not password:
        return {"Message": "Missing username or password.."}
    if Parser.query.filter_by(username=username).first():
        return {"MEssage": "Username is already exist.."}

    new_user = Parser(username=username, password=password)
    db.session.add(new_user)
    db.session.commit()
    return {"Message": " New user is Successfully created..."}"""


@app.route("/login", methods=["POST"])
def login():
    username = request.json.get("username", None)
    password = request.json.get("password", None)

    if username == "username" and password == "password":
        access_token = create_access_token(identity=username)
        return jsonify(access_token=access_token), 200

    else:
        return jsonify({"MEssage": "Invalid username or password"}), 401


# add data to excel file..
@app.route("/add", methods=["POST"])
@jwt_required()
def add_data():
    if request.method == "POST":
        e_data = request.files["Mydata"]
        Mydata = load_workbook(e_data)
        Newdata = Mydata.active

        for i in Newdata.iter_rows(min_row=2, values_only=True):
            data = Parser(name=i[0], age=i[1], location=i[2])
            db.session.add(data)
            db.session.commit()
    return "Message : Data  retrieved  Successfully...."


#  get data from db...


@app.route("/get", methods=["GET"])
@jwt_required()
def fetch_data():
    all_posts = Parser.query.all()
    result = Parser_schemas.dump(all_posts)
    return jsonify(result)


# update data
@app.route(
    "/update/<int:id>",
    methods=["PUT", "POST"],
)
@jwt_required()
def update_data(id):
    post = Parser.query.get(id)
    name = request.json["name"]
    age = request.json["age"]
    location = request.json["location"]

    post.name = name
    post.age = age
    post.location = location

    db.session.add(post)
    db.session.commit()
    return Parser_schema.jsonify(post)


# Delete data from db...
@app.route("/delete/<int:id>", methods=["DELETE"])
@jwt_required()
def delete_db(id):
    delete = Parser.query.get(id)
    db.session.delete(delete)
    db.session.commit()
    return Parser_schema.jsonify(delete)


if __name__ == "__main__":
    app.run(debug=True, port=1998)
